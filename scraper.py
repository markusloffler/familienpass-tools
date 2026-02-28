"""
Familienpass Event Scraper

Scrapes event data from Munich Familienpass website and exports to Excel
"""

import os
import sys
import time
from typing import List, Dict, Tuple
from urllib.parse import urljoin

from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from config import BASE_URL, TOTAL_PAGES, DELAY_BETWEEN_PAGES, DELAY_BETWEEN_EVENTS, CONTINUATION_LINK_MARKER
from utils import make_request_with_retry, clean_text, extract_field_by_header


def parse_listing_page(html_content: str) -> List[Tuple[str, str]]:
    """
    Parse a single listing page and extract event URLs and dates

    Args:
        html_content: HTML content of the listing page

    Returns:
        List of tuples containing (event_url, date)
    """
    soup = BeautifulSoup(html_content, 'html.parser')
    events = []

    # Find ALL tables with events (one per month)
    tables = soup.find_all('table', class_='events-table')

    if not tables:
        print("  Warning: No tables found on page")
        return events

    for table in tables:
        # Skip header row, iterate through event rows
        rows = table.find_all('tr')[1:]  # Skip first row (headers)

        for row in rows:
            cells = row.find_all('td')
            if len(cells) >= 2:
                # Extract date from first column
                date_text = clean_text(cells[0].get_text())

                # Extract event URL from second column
                link = cells[1].find('a')
                if link and link.get('href'):
                    event_url = link.get('href')

                    # Convert relative URL to absolute if needed
                    if not event_url.startswith('http'):
                        event_url = urljoin('https://veranstaltungen.muenchen.de', event_url)

                    events.append((event_url, date_text))

    return events


def get_all_event_urls(base_url: str, total_pages: int = TOTAL_PAGES) -> List[Tuple[str, str]]:
    """
    Iterate through all listing pages and collect event URLs

    Args:
        base_url: Base URL of the listing page
        total_pages: Number of pages to scrape

    Returns:
        List of tuples containing (event_url, date)
    """
    all_events = []

    for page_num in range(1, total_pages + 1):
        # Construct URL
        if page_num == 1:
            url = base_url
        else:
            url = f"{base_url}?pno={page_num}"

        print(f"Scraping page {page_num}/{total_pages}...")

        try:
            # Fetch page with retry logic
            response = make_request_with_retry(url)

            # Parse and extract event URLs
            events = parse_listing_page(response.text)
            all_events.extend(events)

            print(f"  Found {len(events)} events on this page")

        except Exception as e:
            print(f"  ERROR: Failed to scrape page {page_num}: {e}")

        # Rate limiting - be respectful
        if page_num < total_pages:
            time.sleep(DELAY_BETWEEN_PAGES)

    return all_events


def parse_event_page(html_content: str, event_url: str) -> Dict[str, str]:
    """
    Parse individual event page and extract all 8 fields

    Args:
        html_content: HTML content of the event page
        event_url: URL of the event (for reference)

    Returns:
        Dictionary with event data
    """
    soup = BeautifulSoup(html_content, 'html.parser')

    event_data = {
        'name': '',
        'age': '',
        'place': '',
        'date': '',
        'time': '',
        'sign_up_date': '',
        'event_url': event_url,
        'description': ''
    }

    try:
        # 1. Extract Name from <h2>
        h2 = soup.find('h2')
        if h2:
            event_data['name'] = clean_text(h2.get_text())

        # 2. Extract Age (Alter)
        event_data['age'] = extract_field_by_header(soup, 'Alter')

        # 3. Extract Place (Treffpunkt)
        event_data['place'] = extract_field_by_header(soup, 'Treffpunkt')

        # 4. Extract Date (Datum)
        event_data['date'] = extract_field_by_header(soup, 'Datum')

        # 5. Extract Time (Uhrzeit)
        event_data['time'] = extract_field_by_header(soup, 'Uhrzeit')

        # 6. Extract Sign-up information (Verlosungszeitraum or direct registration)
        sign_up = ''

        # First, try to find "Anmeldebeginn" section with "Verlosungszeitraum: date range"
        anmeldebeginn = extract_field_by_header(soup, 'Anmeldebeginn')
        if anmeldebeginn and 'Verlosungszeitraum' in anmeldebeginn:
            # Extract just the date part after "Verlosungszeitraum:"
            if ':' in anmeldebeginn:
                sign_up = anmeldebeginn.split(':', 1)[1].strip()
            else:
                sign_up = anmeldebeginn

            # Replace " bis " with " - "
            sign_up = sign_up.replace(' bis ', ' - ')

        # If no lottery period, check for direct registration with organizer
        if not sign_up:
            anmeldung = extract_field_by_header(soup, 'Anmeldung')
            if anmeldung and 'direkt beim Veranstalter' in anmeldung:
                sign_up = 'Die Anmeldung erfolgt direkt beim Veranstalter.'

        event_data['sign_up_date'] = sign_up if sign_up else 'Not specified'

        # 8. Extract Description (main content paragraph after title, before structured sections)
        description = ''

        # Find the first substantial paragraph after h2 but before first h3
        if h2:
            # Get all siblings after h2
            current = h2.find_next_sibling()
            while current:
                # Stop if we hit a structured section (h3 header)
                if current.name == 'h3':
                    break

                # Check if it's a paragraph with substantial content
                if current.name == 'p':
                    text = clean_text(current.get_text())
                    if len(text) > 50:  # Substantial content
                        description = text
                        break

                current = current.find_next_sibling()

        # Fallback: try to find any substantial paragraph
        if not description:
            paragraphs = soup.find_all('p')
            for p in paragraphs:
                text = clean_text(p.get_text())
                if len(text) > 50:  # Substantial content
                    description = text
                    break

        event_data['description'] = description if description else 'No description available'

    except Exception as e:
        print(f"    Warning: Error parsing event page: {e}")

    return event_data


def scrape_event_details(event_url: str, listing_date: str = '') -> Dict[str, str]:
    """
    Scrape individual event page and extract all 8 fields

    Args:
        event_url: URL of the event to scrape
        listing_date: Date extracted from listing page (fallback)

    Returns:
        Dictionary with event data
    """
    print(f"  Scraping: {event_url}")

    try:
        # Fetch page with retry
        response = make_request_with_retry(event_url)

        # Parse and extract
        event_data = parse_event_page(response.text, event_url)

        # Use listing date as fallback if detail page date missing
        if not event_data['date'] and listing_date:
            event_data['date'] = listing_date

        return event_data

    except Exception as e:
        print(f"  ERROR scraping {event_url}: {e}")
        # Return partial data with error indicator
        return {
            'name': f'ERROR: {event_url}',
            'age': '',
            'place': '',
            'date': listing_date if listing_date else '',
            'time': '',
            'sign_up_date': '',
            'event_url': event_url,
            'description': f'Scraping failed: {str(e)}'
        }


def load_existing_selections(filepath: str) -> dict:
    """Load existing selections from Excel file, mapping event name to selection value."""
    selections = {}
    if not os.path.exists(filepath):
        return selections

    try:
        wb = load_workbook(filepath)
        ws = wb.active
        for row_num in range(2, ws.max_row + 1):
            selected = ws.cell(row=row_num, column=1).value  # Column A
            event_name = ws.cell(row=row_num, column=2).value  # Column B
            if event_name and selected and str(selected).strip():
                selections[event_name] = selected
        wb.close()
    except Exception as e:
        print(f"Warning: Could not load existing selections: {e}")

    return selections


def group_events_by_name(events_data: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """
    Group events by name and create continuation rows for multi-date events.

    For events with multiple dates, the first occurrence has all information,
    and subsequent occurrences only have date/time/sign-up period.

    Args:
        events_data: List of event dictionaries

    Returns:
        List of event dictionaries with continuation rows marked
    """
    from collections import OrderedDict

    # Group events by name, preserving order of first occurrence
    grouped = OrderedDict()
    for event in events_data:
        name = event['name']
        if name not in grouped:
            grouped[name] = []
        grouped[name].append(event)

    def parse_date_for_sort(date_str: str) -> tuple:
        """Convert DD.MM.YYYY to sortable tuple (YYYY, MM, DD)."""
        try:
            parts = date_str.split('.')
            if len(parts) == 3:
                return (parts[2], parts[1], parts[0])
        except (AttributeError, IndexError):
            pass
        return ('9999', '99', '99')  # Unknown dates sort last

    # Build result with continuation rows
    result = []
    for name, occurrences in grouped.items():
        # Sort occurrences by date for consistent ordering
        occurrences.sort(key=lambda e: parse_date_for_sort(e.get('date', '')))

        for idx, event in enumerate(occurrences):
            if idx == 0:
                # First occurrence: include all info
                event['is_continuation'] = False
                result.append(event)
            else:
                # Subsequent occurrences: only date/time/sign-up period
                continuation = {
                    'name': name,  # Keep name for selection lookup
                    'event_url': event['event_url'],
                    'description': '',
                    'age': '',
                    'place': '',
                    'date': event['date'],
                    'time': event['time'],
                    'sign_up_date': event['sign_up_date'],
                    'is_continuation': True
                }
                result.append(continuation)

    return result


def save_to_excel(events_data: List[Dict[str, str]], output_path: str, selections: dict = None) -> None:
    """
    Create Excel file with formatted event data

    Args:
        events_data: List of event dictionaries
        output_path: Path to save the Excel file
        selections: Dict mapping event name to selection value
    """
    if selections is None:
        selections = {}

    # Group events by name (multi-date events get continuation rows)
    grouped_events = group_events_by_name(events_data)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Familienpass Events"

    # Define column headers with "Selected" as first column
    headers = [
        'Selected',
        'Event Name',
        'Description',
        'Age (Alter)',
        'Place (Treffpunkt)',
        'Date',
        'Time',
        'Sign-Up Period (Verlosungszeitraum)'
    ]

    # Write headers with formatting
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, size=12, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data rows
    for row_num, event in enumerate(grouped_events, 2):
        is_continuation = event.get('is_continuation', False)

        if is_continuation:
            # Continuation row: only date/time/sign-up period (no selection)
            ws.cell(row=row_num, column=1).alignment = Alignment(wrap_text=True, vertical='top')

            # Column 2: "Link" as hyperlink to this occurrence's event URL
            link_cell = ws.cell(row=row_num, column=2, value=CONTINUATION_LINK_MARKER)
            link_cell.hyperlink = event['event_url']
            link_cell.font = Font(color='0563C1', underline='single')
            link_cell.alignment = Alignment(wrap_text=True, vertical='top')

            ws.cell(row=row_num, column=3).alignment = Alignment(wrap_text=True, vertical='top')
            ws.cell(row=row_num, column=4).alignment = Alignment(wrap_text=True, vertical='top')
            ws.cell(row=row_num, column=5).alignment = Alignment(wrap_text=True, vertical='top')
        else:
            # First row: full event info with selection
            selection_value = selections.get(event['name'], '')
            selected_cell = ws.cell(row=row_num, column=1, value=selection_value)
            selected_cell.alignment = Alignment(wrap_text=True, vertical='top')

            # Column 2: Event Name (as hyperlink)
            name_cell = ws.cell(row=row_num, column=2, value=event['name'])
            name_cell.hyperlink = event['event_url']
            name_cell.font = Font(color='0563C1', underline='single')
            name_cell.alignment = Alignment(wrap_text=True, vertical='top')

            # Column 3: Description
            desc_cell = ws.cell(row=row_num, column=3, value=event['description'])
            desc_cell.alignment = Alignment(wrap_text=True, vertical='top')

            # Column 4: Age
            age_cell = ws.cell(row=row_num, column=4, value=event['age'])
            age_cell.alignment = Alignment(wrap_text=True, vertical='top')

            # Column 5: Place
            place_cell = ws.cell(row=row_num, column=5, value=event['place'])
            place_cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Date/Time/Sign-up always filled (for both first and continuation rows)
        date_cell = ws.cell(row=row_num, column=6, value=event['date'])
        date_cell.alignment = Alignment(wrap_text=True, vertical='top')

        time_cell = ws.cell(row=row_num, column=7, value=event['time'])
        time_cell.alignment = Alignment(wrap_text=True, vertical='top')

        signup_cell = ws.cell(row=row_num, column=8, value=event['sign_up_date'])
        signup_cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Set column widths (with Selected column first)
    column_widths = {
        'A': 10.0,         # Selected
        'B': 37.33203125,  # Event Name
        'C': 56.33203125,  # Description
        'D': 12.5,         # Age
        'E': 21.83203125,  # Place
        'F': 10.1640625,   # Date
        'G': 14.0,         # Time
        'H': 22.33203125   # Sign-Up Period
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Set zoom level to 150%
    ws.sheet_view.zoomScale = 150

    # Create output directory if needed
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Save workbook
    wb.save(output_path)
    print(f"Excel file saved: {output_path}")


def main():
    """
    Main entry point - orchestrates entire scraping process
    """
    try:
        print("=" * 50)
        print("Familienpass Event Scraper")
        print("=" * 50)
        print()

        # Step 1: Get all event URLs
        print("Step 1: Collecting event URLs from listing pages...")
        event_urls = get_all_event_urls(BASE_URL, total_pages=TOTAL_PAGES)
        print(f"\nFound {len(event_urls)} events total\n")

        if not event_urls:
            print("ERROR: No events found. Exiting.")
            sys.exit(1)

        # Step 2: Scrape each event
        print("Step 2: Scraping individual event details...")
        events_data = []

        for idx, (url, listing_date) in enumerate(event_urls, 1):
            print(f"Event {idx}/{len(event_urls)}")
            event_data = scrape_event_details(url, listing_date)
            events_data.append(event_data)

            # Rate limiting
            if idx < len(event_urls):
                time.sleep(DELAY_BETWEEN_EVENTS)

        # Step 3: Save to Excel
        print("\nStep 3: Saving to Excel...")
        output_path = 'output/familienpass_events.xlsx'
        existing_selections = load_existing_selections(output_path)
        if existing_selections:
            print(f"Loaded {len(existing_selections)} existing selections")
        save_to_excel(events_data, output_path, existing_selections)

        # Count unique events for reporting
        unique_names = set(e['name'] for e in events_data)

        print()
        print("=" * 50)
        print(f"Success! Scraped {len(events_data)} event occurrences ({len(unique_names)} unique events)")
        print(f"Output saved to: {output_path}")
        print("=" * 50)

    except KeyboardInterrupt:
        print("\n\nScraping interrupted by user. Exiting...")
        sys.exit(0)

    except Exception as e:
        print(f"\nFATAL ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
