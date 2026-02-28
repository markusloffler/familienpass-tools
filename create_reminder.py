"""
Create Apple Reminders from Familienpass Excel

Reads the Excel file and creates Apple Reminders for selected events.
Requires macOS with Reminders access granted to the terminal.
"""

import argparse
import os
from datetime import datetime
from openpyxl import load_workbook

from create_calendar import parse_date_range
from config import CONTINUATION_LINK_MARKER


REMINDER_LIST_NAME = "Familienpass"


def read_selected_events(excel_path):
    """
    Read Excel file and extract selected events with their properties.

    Returns:
        List of dicts with keys: title, url, sign_up_period, date, time, start_date
    """
    wb = load_workbook(excel_path)
    ws = wb.active

    events = []
    current_selected_event = None
    current_event_url = None

    for row_num in range(2, ws.max_row + 1):
        selected_cell = ws.cell(row=row_num, column=1)
        event_name_cell = ws.cell(row=row_num, column=2)
        date_cell = ws.cell(row=row_num, column=6)
        time_cell = ws.cell(row=row_num, column=7)
        sign_up_cell = ws.cell(row=row_num, column=8)

        event_name = event_name_cell.value
        is_continuation = not event_name or str(event_name).strip() in ('', CONTINUATION_LINK_MARKER)

        if not is_continuation:
            selected = selected_cell.value
            if selected and str(selected).strip() != '':
                current_selected_event = event_name
                current_event_url = (
                    event_name_cell.hyperlink.target if event_name_cell.hyperlink else None
                )
            else:
                current_selected_event = None
                current_event_url = None

        if not current_selected_event:
            continue

        # For continuation rows, use this occurrence's URL; otherwise use the main row's URL
        row_url = (
            event_name_cell.hyperlink.target
            if is_continuation and event_name_cell.hyperlink
            else current_event_url
        )

        sign_up_period = sign_up_cell.value
        date_range = parse_date_range(sign_up_period)
        if not date_range:
            if not is_continuation:
                print(f"  Skipping '{current_selected_event}': No valid sign-up date")
            continue

        start_date, _ = date_range
        event_date = date_cell.value or ''
        event_time = time_cell.value or ''

        events.append({
            'title': f"Anmeldung Familienpass: {current_selected_event}",
            'url': row_url,
            'sign_up_period': sign_up_period,
            'date': str(event_date).strip(),
            'time': str(event_time).strip(),
            'start_date': start_date,
        })

    return events


def build_notes(event):
    """Build reminder notes string from event properties."""
    lines = []
    if event['url']:
        lines.append(event['url'])
    lines.append(f"Anmeldezeitraum: {event['sign_up_period']}")
    if event['date']:
        date_line = f"Datum: {event['date']}"
        if event['time']:
            date_line += f", {event['time']}"
        lines.append(date_line)
    return '\n'.join(lines)


def dry_run(events):
    """Print all reminders that would be created."""
    print(f"\nDry run - {len(events)} reminder(s) would be created:\n")

    for i, event in enumerate(events, 1):
        print(f"--- Reminder {i} ---")
        notes = build_notes(event)
        indented_notes = notes.replace('\n', '\n            ')
        print(f"  Title:    {event['title']}")
        print(f"  Due date: {event['start_date'].strftime('%d.%m.%Y')}")
        print(f"  URL:      {event['url'] or '(none)'}")
        print(f"  Notes:    {indented_notes}")
        print()


def ensure_reminder_list(remind):
    """Find the Familienpass reminder list, or create it via EventKit."""
    try:
        return remind.calendars.get(REMINDER_LIST_NAME)
    except ValueError:
        pass

    print(f"  Reminder list '{REMINDER_LIST_NAME}' not found. Creating it...")
    import EventKit
    store = remind._event_store
    new_cal = EventKit.EKCalendar.calendarForEntityType_eventStore_(
        EventKit.EKEntityTypeReminder, store
    )
    new_cal.setTitle_(REMINDER_LIST_NAME)
    new_cal.setSource_(store.defaultCalendarForNewReminders().source())
    error = None
    store.saveCalendar_commit_error_(new_cal, True, error)
    return remind.calendars.get(REMINDER_LIST_NAME)


def create_reminders(events):
    """Create Apple Reminders for the given events."""
    from pyremindkit import RemindKit

    remind = RemindKit()
    target_calendar = ensure_reminder_list(remind)

    created = 0
    for event in events:
        notes = build_notes(event)
        target_calendar.create_reminder(
            title=event['title'],
            due_date=event['start_date'],
            notes=notes,
        )
        created += 1
        print(f"  Created: {event['title']} (due {event['start_date'].strftime('%d.%m.%Y')})")

    print(f"\n{created} reminder(s) created in list '{REMINDER_LIST_NAME}'.")


def main():
    parser = argparse.ArgumentParser(description='Create Apple Reminders from Familienpass Excel')
    parser.add_argument(
        '--dry-run', action='store_true',
        help='Only print reminders that would be created, without actually creating them'
    )
    args = parser.parse_args()

    excel_path = 'output/familienpass_events.xlsx'

    print("=" * 50)
    print("Familienpass Reminder Creator")
    print("=" * 50)
    print()

    if not os.path.exists(excel_path):
        print(f"ERROR: Excel file not found: {excel_path}")
        print("Please run the scraper first to generate the Excel file.")
        return

    events = read_selected_events(excel_path)

    if not events:
        print("No events to create reminders for.")
        print("Please mark events in the 'Selected' column of the Excel file.")
        return

    if args.dry_run:
        dry_run(events)
    else:
        create_reminders(events)

    print()
    print("=" * 50)


if __name__ == '__main__':
    main()
