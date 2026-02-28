"""
Create Calendar Events from Familienpass Excel

Reads the Excel file and creates .ics calendar events for selected events
"""

import hashlib
import os
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook
from icalendar import Calendar, Event, Alarm
from config import CONTINUATION_LINK_MARKER


def parse_date_range(date_str):
    """
    Parse date string in format "DD.MM.YYYY - DD.MM.YYYY" or "DD.MM.YYYY"

    Args:
        date_str: Date string from sign-up period column

    Returns:
        Tuple of (start_date, end_date) as datetime objects, or None if parsing fails
    """
    if not date_str or date_str == 'Not specified':
        return None

    # Check if it's the direct registration message
    if 'direkt beim Veranstalter' in date_str:
        return None

    # Remove any extra whitespace
    date_str = date_str.strip()

    def parse_date(s):
        """Parse the first DD.MM.YYYY date from a string, ignoring trailing text."""
        m = re.match(r'\d{2}\.\d{2}\.\d{4}', s.strip())
        if m:
            try:
                return datetime.strptime(m.group(), '%d.%m.%Y')
            except ValueError:
                pass
        return None

    # Check if it's a range
    if ' - ' in date_str:
        parts = date_str.split(' - ', 1)
        start_date = parse_date(parts[0])
        end_date = parse_date(parts[1])
        if start_date and end_date:
            return (start_date, end_date)
        return None
    else:
        date = parse_date(date_str)
        return (date, date) if date else None


def create_calendar_events(excel_path, output_path):
    """
    Read Excel file and create .ics calendar file for selected events

    Args:
        excel_path: Path to the Excel file
        output_path: Path where .ics file should be saved
    """
    # Load workbook
    wb = load_workbook(excel_path)
    ws = wb.active

    # Create calendar
    cal = Calendar()
    cal.add('prodid', '-//Familienpass Events//muenchen.de//')
    cal.add('version', '2.0')

    events_created = 0
    events_skipped = 0

    # Track currently selected event for handling continuation rows
    current_selected_event = None
    current_event_url = None

    # Iterate through rows (skip header)
    for row_num in range(2, ws.max_row + 1):
        # Column indices (1-based for cells):
        # A (1): Selected
        # B (2): Event Name
        # C (3): Description
        # D (4): Age
        # E (5): Place
        # F (6): Date
        # G (7): Time
        # H (8): Sign-Up Period

        selected_cell = ws.cell(row=row_num, column=1)
        event_name_cell = ws.cell(row=row_num, column=2)
        sign_up_cell = ws.cell(row=row_num, column=8)

        selected = selected_cell.value
        event_name = event_name_cell.value
        sign_up_period = sign_up_cell.value

        # Check if this is a main event row (has event name) or continuation row (↗ or blank)
        is_continuation = not event_name or str(event_name).strip() in ('', CONTINUATION_LINK_MARKER)

        if not is_continuation:
            # Main event row - check if selected
            if selected and str(selected).strip() != '':
                current_selected_event = event_name
                current_event_url = event_name_cell.hyperlink.target if event_name_cell.hyperlink else None
            else:
                current_selected_event = None
                current_event_url = None

        # Skip if current event is not selected
        if not current_selected_event:
            continue

        # For continuation rows, use this occurrence's URL; otherwise use the main row's URL
        row_url = (
            event_name_cell.hyperlink.target
            if is_continuation and event_name_cell.hyperlink
            else current_event_url
        )

        # Parse sign-up period
        date_range = parse_date_range(sign_up_period)

        if not date_range:
            if not is_continuation:
                print(f"  Skipping '{current_selected_event}': No valid sign-up date")
            events_skipped += 1
            continue

        start_date, end_date = date_range

        # Create event
        event = Event()
        event.add('summary', f'Anmeldung Familienpass: {current_selected_event}')
        event.add('dtstart', start_date.date())
        # Add 1 day to end_date because DTEND is exclusive in iCalendar format
        event.add('dtend', (end_date + timedelta(days=1)).date())
        event.add('dtstamp', datetime.now())

        # Create stable UID from event name + date range to prevent duplicates on re-import
        uid_source = f"{current_selected_event}-{start_date.isoformat()}-{end_date.isoformat()}"
        uid = hashlib.sha256(uid_source.encode()).hexdigest()[:32] + "@familienpass"
        event.add('uid', uid)

        # Add URL if available
        if row_url:
            event.add('url', row_url)

        # Add alarm on the day of the event (at start)
        alarm = Alarm()
        alarm.add('action', 'DISPLAY')
        alarm.add('trigger', timedelta(0))  # 0 = at event start
        alarm.add('description', f'Anmeldung Familienpass: {current_selected_event}')
        event.add_component(alarm)

        # Add to calendar
        cal.add_component(event)
        events_created += 1
        print(f"  Created: Anmeldung Familienpass: {current_selected_event} ({start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')})")

    # Save calendar file
    if events_created > 0:
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
        with open(output_path, 'wb') as f:
            f.write(cal.to_ical())

        print(f"\nCalendar file created: {output_path}")
        print(f"Total events: {events_created}")
        if events_skipped > 0:
            print(f"Skipped (no date): {events_skipped}")
    else:
        print("\nNo events to create.")
        print("Please mark events in the 'Selected' column of the Excel file.")


def main():
    """
    Main entry point
    """
    excel_path = 'output/familienpass_events.xlsx'
    output_path = 'output/familienpass_calendar.ics'

    print("=" * 50)
    print("Familienpass Calendar Event Creator")
    print("=" * 50)
    print()

    if not os.path.exists(excel_path):
        print(f"ERROR: Excel file not found: {excel_path}")
        print("Please run the scraper first to generate the Excel file.")
        return

    create_calendar_events(excel_path, output_path)

    print()
    print("=" * 50)


if __name__ == '__main__':
    main()
